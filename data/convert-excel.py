#!/usr/bin/env python3
"""
Convert the wedding guest Excel file to guests.json for the RSVP system.
Run: python3 data/convert-excel.py
"""

import json
import os
import openpyxl
import jellyfish

EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', 'Wedding - List_and_Pre-Wed -vWebsite.xlsx')
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), 'guests.json')

def metaphone_parts(name):
    """Generate metaphone keys for first and last name parts."""
    parts = name.strip().split()
    first = jellyfish.metaphone(parts[0]) if parts else ''
    last = jellyfish.metaphone(parts[-1]) if len(parts) > 1 else ''
    return first, last

def safe_int(val):
    """Convert a value to int, returning 0 for non-numeric values."""
    try:
        return int(val)
    except (ValueError, TypeError):
        return 0

def extract_two_column_sheet(ws, header_row, data_start_row):
    """Extract guests from sheets with side-by-side two-column layout."""
    guests = []
    for row in ws.iter_rows(min_row=data_start_row, values_only=False):
        # Left group: columns C(2), D(3), E(4)
        name_left = row[2].value
        plus_one_left = row[3].value
        pre_wed_left = row[4].value

        if name_left and str(name_left).strip() and str(name_left).strip().lower() != 'total':
            name = str(name_left).strip()
            po = safe_int(plus_one_left)
            guests.append({
                'name': name,
                'plus_one_allowed': po >= 1,
                'pre_wedding_invited': str(pre_wed_left or '').strip().upper() == 'Y',
                'headcount': 1 + po,
            })

        # Right group: columns G(6), H(7), I(8)
        if len(row) > 8:
            name_right = row[6].value
            plus_one_right = row[7].value
            pre_wed_right = row[8].value

            if name_right and str(name_right).strip() and str(name_right).strip().lower() != 'total':
                name = str(name_right).strip()
                po = safe_int(plus_one_right)
                guests.append({
                    'name': name,
                    'plus_one_allowed': po >= 1,
                    'pre_wedding_invited': str(pre_wed_right or '').strip().upper() == 'Y',
                    'headcount': 1 + po,
                })

    return guests

def extract_parents_sheet(ws):
    """Extract guests from the Parents sheet (single column, couples)."""
    guests = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        headcount = row[1]
        name = row[2]
        if name and str(name).strip():
            guests.append({
                'name': str(name).strip(),
                'plus_one_allowed': False,
                'pre_wedding_invited': True,
                'headcount': int(headcount) if headcount else 2,
            })
    return guests

def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    all_guests = []

    # Sheet "Corine": header at row 3 (0-indexed 2), data starts at row 5 (0-indexed 4)
    if 'Corine' in wb.sheetnames:
        ws = wb['Corine']
        all_guests.extend(extract_two_column_sheet(ws, header_row=3, data_start_row=5))

    # Sheet "Isaac": header at row 2 (0-indexed 1), data starts at row 4 (0-indexed 3)
    if 'Isaac' in wb.sheetnames:
        ws = wb['Isaac']
        all_guests.extend(extract_two_column_sheet(ws, header_row=2, data_start_row=4))

    # Sheet "Parents 1": data starts at row 2
    if 'Parents 1' in wb.sheetnames:
        ws = wb['Parents 1']
        all_guests.extend(extract_parents_sheet(ws))

    # Deduplicate by name_lower
    seen = set()
    unique_guests = []
    for g in all_guests:
        key = g['name'].lower()
        if key not in seen:
            seen.add(key)
            unique_guests.append(g)

    # Add id, name_lower, metaphone keys
    output = []
    for i, g in enumerate(unique_guests, start=1):
        first_mp, last_mp = metaphone_parts(g['name'])
        output.append({
            'id': i,
            'name': g['name'],
            'name_lower': g['name'].lower(),
            'first_name_metaphone': first_mp,
            'last_name_metaphone': last_mp,
            'plus_one_allowed': g['plus_one_allowed'],
            'pre_wedding_invited': g['pre_wedding_invited'],
            'headcount': g['headcount'],
        })

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    print(f"Wrote {len(output)} guests to {OUTPUT_PATH}")

if __name__ == '__main__':
    main()
